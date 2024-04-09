import os
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from PyPDF2 import PdfReader, PdfWriter
import logging

# Function to combine PDFs based on input_folder, output_folder, and sheet
def combine_pdfs(input_folder, output_folder, sheet):
    """
    Function to combine multiple PDFs from the input folder into a single PDF for each release in the output folder based on the provided sheet data.
    
    Args:
        input_folder (str): The path to the input folder containing the PDF files to be combined.
        output_folder (str): The path to the output folder where the combined PDFs will be saved.
        sheet (pandas.DataFrame): The sheet data containing the information about the PDFs to be combined.
    """
    pdf_writer = PdfWriter()
    last_release = None
    added_pdfs = []
    unique_ffid_dates = set()  # Set to store unique {FFID}{WorkDate} combinations
    missing_files = set()  # Set to store missing files

    # Iterate through rows in the sheet
    for release, ffid, _, description, work_date, _, _ in sheet.iter_rows(min_row=2, max_col=7, values_only=True):
        if ffid and work_date:
            # Check if the release has changed
            if last_release is not None and release != last_release:
                output_file_path = os.path.join(output_folder, f"{last_release}.pdf").replace("\\", "/")
                with open(output_file_path, "wb") as output_file:
                    pdf_writer.write(output_file)
                pdf_writer = PdfWriter()
                added_pdfs = []

            output_file_name = f"{release}.pdf"
            output_file_path = os.path.join(output_folder, output_file_name).replace("\\", "/")


            # Decide file name based on description
            if "CXL" in description:
                input_file_name = f"{ffid}  Duke Energy.pdf"
            else:
                input_file_name = f"{ffid} {work_date.strftime('%Y-%m-%d')} Duke Energy.pdf"

            input_file_path = os.path.join(input_folder, input_file_name).replace("\\", "/")

            ffid_date_key = f"{ffid}_{work_date.strftime('%Y-%m-%d')}"

            # Check if {FFID}{WorkDate} combination is unique
            if ffid_date_key not in unique_ffid_dates:
                print(f"Processing {release}")
                unique_ffid_dates.add(ffid_date_key)  # Add unique {FFID}{WorkDate} combination to the set

                if os.path.exists(input_file_path):
                    added_pdfs.append(input_file_path)
                    pdf_reader = PdfReader(input_file_path)
                    for page in pdf_reader.pages:
                        pdf_writer.add_page(page)
                else:
                    # Add file and associated WO# to a set
                    if "CXL" in description:
                        input_file_name = f"{ffid}  Piedmont Natural Gas Company, Inc..pdf"
                    else:
                        input_file_name = f"{ffid} {work_date.strftime('%Y-%m-%d')} Piedmont Natural Gas Company, Inc..pdf"

                    if os.path.exists(input_file_path):
                        added_pdfs.append(input_file_path)
                        pdf_reader = PdfReader(input_file_path)
                        for page in pdf_reader.pages:
                            pdf_writer.add_page(page)
                    else:
                        missing_files.add(f"{release} - {os.path.basename(input_file_path)}")

                    continue

            last_release = release

    # Process the last release
    if last_release is not None:
        output_file_path = os.path.join(output_folder, f"{last_release}.pdf").replace("\\", "/")
        with open(output_file_path, "wb") as output_file:
            pdf_writer.write(output_file)

    # Create logger and set level to debug
    logger = logging.getLogger()
    logger.setLevel(logging.DEBUG)

    # Create file handler which logs debug messages
    output_log_path = os.path.join(output_folder, "debug.log")
    fh = logging.FileHandler(output_log_path)
    fh.setLevel(logging.DEBUG)

    # Add the handlers to logger
    logger.addHandler(fh)

    # Log missing files
    for file in missing_files:
        logger.debug(file)

    print("Done!")
    
def get_input_directory(root, label):
    input_dir = filedialog.askdirectory(title="Select the Input Folder")
    if input_dir:
        root.input_dir = input_dir
        label.config(text=f"Input folder selected: {os.path.basename(input_dir)}")

def get_output_directory(root, label):
    output_dir = filedialog.askdirectory(title="Select the Output Folder")
    if output_dir:
        root.output_dir = output_dir
        label.config(text=f"Output folder selected: {os.path.basename(output_dir)}")

def get_excel_file(root, label):
    excel_file = filedialog.askopenfilename(title="Select the Excel Workbook", 
                                            filetypes = [("Excel files", "*.xlsx"), ("All files", "*.*")])
    if excel_file:
        root.excel_file = excel_file
        label.config(text=f"Excel file selected: {os.path.basename(excel_file)}")

def start_combining(root, completion_label):
    if hasattr(root, 'input_dir') and hasattr(root, 'output_dir') and hasattr(root, 'excel_file'):
        workbook = load_workbook(root.excel_file)
        sheet = workbook["TimesheetCombiner"]
        combine_pdfs(root.input_dir, root.output_dir, sheet)
        completion_label.config(text="Combining Completed. Check the output folder for combined PDFs.")
    else:
        messagebox.showerror("Error", "Make sure to select an input folder, an output folder, and an Excel file.")

def main():
    root = tk.Tk()
    root.title("Timesheet Combiner")
    root.geometry("500x300")

    # Instructions and selection buttons
    label_instructions = tk.Label(root, text="1. Select the input folder with PDFs.\n2. Select the output folder for combined PDFs.\n3. Select the Excel workbook for PDF details.")
    label_instructions.pack(pady=10)

    # Selection and status labels for input folder, output folder, and Excel file
    input_folder_label = tk.Label(root, text="Input Folder: Not selected")
    output_folder_label = tk.Label(root, text="Output Folder: Not selected")
    excel_file_label = tk.Label(root, text="Excel File: Not selected")
    
    input_folder_label.pack()
    tk.Button(root, text="Select Input Folder", command=lambda: get_input_directory(root, input_folder_label)).pack()
    output_folder_label.pack()
    tk.Button(root, text="Select Output Folder", command=lambda: get_output_directory(root, output_folder_label)).pack()
    excel_file_label.pack()
    tk.Button(root, text="Select Excel File", command=lambda: get_excel_file(root, excel_file_label)).pack()

    start_button = tk.Button(root, text="Start Combining", command=lambda: start_combining(root, completion_label))
    start_button.pack(pady=15)

    completion_label = tk.Label(root, text="")
    completion_label.pack()

    root.mainloop()

if __name__ == "__main__":
    main()
