import os
import logging
from datetime import datetime
from tkinter import Tk, filedialog
from openpyxl import load_workbook
from PyPDF2 import PdfReader, PdfWriter

def combine_timesheets(input_folder, output_folder, sheet):
    """
    Combine multiple timesheets into a single PDF file for each unique TC number.

    Args:
        input_folder (str): The path to the folder containing the input timesheet files.
        output_folder (str): The path to the folder where the combined PDF files will be saved.
        sheet (Sheet): The Excel sheet containing the timesheet data.

    Returns:
        None
    """
    unique_tc_numbers = set()  # Set to store unique TC numbers
    tc_records = {}  # Dictionary to store records for each TC number
    missing_files = set()  # Set to store missing files

    # Iterate through rows in the sheet, collect unique TC numbers and their corresponding records
    for row, (tc_number, ffid, work_date, invoice) in enumerate(sheet.iter_rows(min_row=2, max_col=4, values_only=True), start=2):
        # Convert work_date to a datetime object and format it
        formatted_work_date = work_date.strftime("%Y-%m-%d")
        
        # Create a unique record {ffid}{WorkDate}
        unique_id = f"{ffid}_{formatted_work_date}"
        
        # Add tc_number to the unique_tc_numbers set
        unique_tc_numbers.add(tc_number)
        
        # Add unique_id to the tc_number's records in the tc_records dict
        if tc_number in tc_records:
            tc_records[tc_number].append(unique_id)
        else:
            tc_records[tc_number] = [unique_id]
     
    # Process each unique TC number
    for tc_number in unique_tc_numbers:
        pdf_writer = PdfWriter()  # New PdfWriter for each tc_number

        # Process each record for the current TC number
        for unique_id in tc_records[tc_number]:
            ffid, formatted_work_date = unique_id.split('_')  # Extract ffid and formatted_work_date from the unique_id
            
            # Define the file name and path for the timesheet
            timesheet_file_name = f"{ffid} {formatted_work_date} Verizon - PA.pdf"
            timesheet_file_path = os.path.join(input_folder, timesheet_file_name).replace("\\", "/")        
        
            # Check if the timesheet file exists
            if os.path.exists(timesheet_file_path):
                # If it exists, read the PDF and add its pages to the pdf_writer
                pdf_reader = PdfReader(timesheet_file_path)
                print(f"{tc_number}: {ffid} {formatted_work_date}")
                for page in pdf_reader.pages:
                    pdf_writer.add_page(page)
            else:
                # If it doesn't exist, add it to the missing_files set
                print(f"{timesheet_file_name} is missing")
                missing_files.add(timesheet_file_name)

        # Write the combined PDFs to a file
        output_file_path = os.path.join(output_folder, f"{tc_number}.pdf").replace("\\", "/")
        with open(output_file_path, "wb") as output_file:
            pdf_writer.write(output_file)

    # Create logger and set level to debug
    logger = logging.getLogger()
    logger.setLevel(logging.DEBUG)

    # Create file handler which logs debug messages
    output_log_path = os.path.join(output_folder, "debug.log")
    fh = logging.FileHandler(output_log_path)
    fh.setLevel(logging.DEBUG)

    # Add the handler to logger
    logger.addHandler(fh)

    # Log missing files
    for file in missing_files:
        logger.debug(file)

    print("Done combining timesheets!")

def combine_invoices(invoice_folder, timesheet_folder, output_folder, sheet):
    """
    Combine invoices from the given invoice folder and timesheets from the timesheet folder into a single PDF file.
    
    Args:
        invoice_folder (str): The path to the folder containing the invoice files.
        timesheet_folder (str): The path to the folder containing the timesheet files.
        output_folder (str): The path to the folder where the combined PDF files will be saved.
        sheet (object): The excel sheet object containing the data.

    Returns:
        None

    Raises:
        FileNotFoundError: If the invoice file or timesheet file does not exist.
    """
    # iterate through the rows starting from row 2 (headers are on 1)
    for row, (tc_number, ffid, work_date, invoice) in enumerate(sheet.iter_rows(min_row=2, max_col=4, values_only=True), start=2):
        # PdfWriter for each invoice
        pdf_writer = PdfWriter()

        # Read work_date and format it
        formatted_work_date = work_date.strftime("%Y-%m-%d")

        # Skip if tc_number contains "CXL" in last 3 chars
        #if tc_number[-3:].upper() == "CXL":
            #continue

        # Define the file names for the invoice, timesheet, and output files        
        invoice_file_name = f"{invoice}.pdf"
        timesheet_file_name = f"{tc_number}.pdf"
        output_file_name = f"{invoice}_{tc_number}.pdf"

        # Define the file paths for the invoice, timesheet, and output files
        invoice_file_path = os.path.join(invoice_folder, invoice_file_name).replace("\\", "/")
        timesheet_file_path = os.path.join(timesheet_folder, timesheet_file_name).replace("\\", "/")
        output_file_path = os.path.join(output_folder, output_file_name).replace("\\", "/")

        # Append invoice file to the output file
        if os.path.exists(invoice_file_path):
            pdf_reader = PdfReader(invoice_file_path)
            for page in pdf_reader.pages:
                pdf_writer.add_page(page)

        # Append timesheet file to the output file
        if os.path.exists(timesheet_file_path):
            pdf_reader = PdfReader(timesheet_file_path)
            for page in pdf_reader.pages:
                pdf_writer.add_page(page)

        # Write the combined PDFs to a file
        with open(output_file_path, "wb") as output_file:
            print(f"{tc_number}, {ffid}, {formatted_work_date}, {invoice}")
            pdf_writer.write(output_file)

    print("Done combining invoices!")

# Main function to handle user interaction and call combine_pdfs(in, out, sheet)
def main():
    """
        Function to execute the main logic of the program.
        The function prompts the user to select various folders and files using file dialogs.
        It then loads an Excel file, retrieves a specific sheet, and performs operations on it.
        Finally, it combines timesheets and invoices, and closes the Tkinter instance.

        Parameters:
        None

        Returns:
        None
    """
    root = Tk()
    root.withdraw()  # Hide the main window

    input_folder = filedialog.askdirectory(title="Select the folder with the exported timesheets from PDES.").replace("\\", "/")
    output_folder = filedialog.askdirectory(title="Select the folder you want to move the combined timesheets into.").replace("\\", "/")

    excel_file_path = filedialog.askopenfilename(title="Select the Excel file", filetypes=[("Excel files", "*.xl*")]).replace("\\", "/")
    workbook = load_workbook(excel_file_path)
    
    # Debug line to print all sheet names
    print(workbook.sheetnames)
    
    sheet = workbook["TimesheetCombiner"]

    combine_timesheets(input_folder, output_folder, sheet)
    
    # Set folders for next run
    invoice_folder = filedialog.askdirectory(title="Select the folder with the newly split invoices.").replace("\\", "/")
    timesheet_folder = output_folder
    output_folder = filedialog.askdirectory(title="Select the folder you want to move the combined PDFs into.").replace("\\", "/")

    # Combine the timesheets with the Invoice in the front
    combine_invoices(invoice_folder, timesheet_folder, output_folder, sheet)

    root.destroy()  # Close the Tkinter instance

if __name__ == "__main__":
    main()
