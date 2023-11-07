import os
import logging
from tkinter import Tk, filedialog
from openpyxl import load_workbook
from PyPDF2 import PdfReader, PdfWriter


def combine_pdfs(input_folder, output_folder, sheet):
    """
    Combines multiple PDF files into a single PDF file.

    Args:
        input_folder (str): The path to the folder containing the input PDF files.
        output_folder (str): The path to the folder where the combined PDF file will be saved.
        sheet (object): The sheet object representing the Excel sheet containing the data.

    Returns:
        None
    """
    pdf_writer = PdfWriter()
    last_release = None
    added_pdfs = []
    unique_ffid_dates = set()  # Set to store unique {FFID}{WorkDate} combinations
    missing_files = set()  # Set to store missing files

    # Iterate through rows in the sheet
    for release, ffid, _, description, work_date, _, _ in sheet.iter_rows(min_row=2, max_col=7, values_only=True):
        if ffid and work_date:
            # Check if the release has changed
            if last_release is not None and release != last_release:
                output_file_path = os.path.join(output_folder, f"{last_release}.pdf").replace("\\", "/")
                with open(output_file_path, "wb") as output_file:
                    pdf_writer.write(output_file)
                pdf_writer = PdfWriter()
                added_pdfs = []

            output_file_name = f"{release}.pdf"
            output_file_path = os.path.join(output_folder, output_file_name).replace("\\", "/")


            # Decide file name based on description
            if "CXL" in description:
                input_file_name = f"{ffid}  Duke Energy.pdf"
            else:
                input_file_name = f"{ffid} {work_date.strftime('%Y-%m-%d')} Duke Energy.pdf"

            input_file_path = os.path.join(input_folder, input_file_name).replace("\\", "/")

            ffid_date_key = f"{ffid}_{work_date.strftime('%Y-%m-%d')}"

            # Check if {FFID}{WorkDate} combination is unique
            if ffid_date_key not in unique_ffid_dates:
                print(f"Processing {release}")
                unique_ffid_dates.add(ffid_date_key)  # Add unique {FFID}{WorkDate} combination to the set

                if os.path.exists(input_file_path):
                    added_pdfs.append(input_file_path)
                    pdf_reader = PdfReader(input_file_path)
                    for page in pdf_reader.pages:
                        pdf_writer.add_page(page)
                else:
                    # Add file and associated WO# to a set
                    missing_files.add(f"{release} - {os.path.basename(input_file_path)}")
                    continue

            last_release = release

    # Process the last release
    if last_release is not None:
        output_file_path = os.path.join(output_folder, f"{last_release}.pdf").replace("\\", "/")
        with open(output_file_path, "wb") as output_file:
            pdf_writer.write(output_file)

    # Create logger and set level to debug
    logger = logging.getLogger()
    logger.setLevel(logging.DEBUG)

    # Create file handler which logs debug messages
    debug_log_file_path = os.path.join(output_folder, 'debug.log')
    fh = logging.FileHandler(debug_log_file_path)
    fh.setLevel(logging.DEBUG)

    # Add the handlers to logger
    logger.addHandler(fh)

    # Log missing files
    for file in missing_files:
        logger.debug(file)

    print("Done!")

def main():
    """
        Main function that performs a series of tasks:
        1. Initializes the Tkinter instance and hides the main window.
        2. Asks the user to select a directory and assigns it to `user_selection`.
        3. Constructs the paths for the input and output folders using `user_selection`.
        4. Asks the user to select an Excel file and assigns its path to `excel_file_path`.
        5. Loads the Excel file using `load_workbook` function from the `openpyxl` library.
        6. Prints all sheet names in the workbook as a debug line.
        7. Retrieves the "TimesheetCombiner" sheet from the workbook.
        8. Calls the `combine_pdfs` function with the input folder, output folder, and the retrieved sheet.
        9. Closes the Tkinter instance.
    """
    root = Tk()
    root.withdraw()  # Hide the main window

    user_selection = filedialog.askdirectory(title="Select the weekly folder").replace("\\", "/")
    input_folder = os.path.join(user_selection, "Inputs").replace("\\", "/")
    output_folder = os.path.join(user_selection, "Outputs").replace("\\", "/")

    excel_file_path = filedialog.askopenfilename(title="Select the Excel file", filetypes=[("Excel files", "*.xl*")]).replace("\\", "/")
    workbook = load_workbook(excel_file_path)
    
    # Debug line to print all sheet names
    print(workbook.sheetnames)
    
    sheet = workbook["TimesheetCombiner"]

    combine_pdfs(input_folder, output_folder, sheet)

    root.destroy()  # Close the Tkinter instance

if __name__ == "__main__":
    main()
