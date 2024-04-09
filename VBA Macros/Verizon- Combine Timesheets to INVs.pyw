import os
import logging
import tkinter as tk
from datetime import datetime
from tkinter import Label, Button, filedialog, messagebox
from openpyxl import load_workbook
from PyPDF2 import PdfReader, PdfWriter

def combine_timesheets(input_folder, combined_folder, sheet, logger):
    """
    Combine timesheets from input_folder into a single timesheet in combined_folder.

    Parameters:
    - input_folder: str, the folder containing the input timesheets
    - combined_folder: str, the folder where the combined timesheet will be saved
    - sheet: Excel sheet object, the sheet containing timesheet data
    - logger: logger object, for logging information

    Returns:
    - None
    """

    unique_tc_numbers = set()
    tc_records = {}
    missing_files = set()

    # Iterate through rows in the sheet
    for row, (tc_number, ffid, work_date, invoice, client) in enumerate(sheet.iter_rows(min_row=2, max_col=5, values_only=True), start=2):
        if '/' in str(tc_number):
            logger.warning(f"Invalid TC number '{tc_number}' in row {row} contains '/'. Skipping.")
            continue

        if work_date is not None:
            formatted_work_date = work_date.strftime("%Y-%m-%d")
            unique_id = f"{ffid}_{formatted_work_date}_{client}"
            unique_tc_numbers.add(tc_number)
            tc_records.setdefault(tc_number, []).append(unique_id)
        else:
            logger.warning(f"Skipping row {row} because work_date is None")

    # Process each unique TC number
    for tc_number in unique_tc_numbers:
        pdf_writer = PdfWriter()

        for unique_id in tc_records[tc_number]:
            ffid, formatted_work_date, client = unique_id.split('_')
            timesheet_file_name = f"{ffid} {formatted_work_date} {client}.pdf"
            timesheet_file_path = os.path.join(input_folder, timesheet_file_name)

            if os.path.exists(timesheet_file_path):
                pdf_reader = PdfReader(timesheet_file_path)
                logger.info(f"Adding {timesheet_file_name} to {tc_number}.pdf")
                for page in pdf_reader.pages:
                    pdf_writer.add_page(page)
            else:
                missing_files.add(timesheet_file_name)

        output_file_path = os.path.join(combined_folder, f"{tc_number}.pdf")
        if not os.path.exists(output_file_path):
            with open(output_file_path, "wb") as output_file:
                pdf_writer.write(output_file)
        else:
            logger.info(f"File {output_file_path} already exists. Skipping.")

    for file in missing_files:
        logger.warning(f"Missing file: {file}")

    logger.info("Done combining timesheets!")
    print("Done combining timesheets!")

def combine_invoices(invoice_folder, timesheet_folder, combined_folder, sheet, logger):
    """
    A function that combines invoices from the given folders into a single output file.
    
    Parameters:
    - invoice_folder: the folder path where the invoice files are stored
    - timesheet_folder: the folder path where the timesheet files are stored
    - combined_folder: the folder path where the combined output files will be saved
    - sheet: the Excel sheet containing the data to be processed
    - logger: the logger object to log messages
    
    Returns:
    None
    """
    # iterate through the rows starting from row 2 (headers are on 1)
    for row, (tc_number, ffid, work_date, invoice, client) in enumerate(sheet.iter_rows(min_row=2, max_col=5, values_only=True), start=2):
        # PdfWriter for each invoice
        pdf_writer = PdfWriter()

        if work_date is not None:
            # Convert work_date to a datetime object and format it
            formatted_work_date = work_date.strftime("%Y-%m-%d")
        else:
            # Handle the case where work_date is None
            # You might want to log this or handle it in some way
            logger.warning(f"Skipping row {row} because work_date is None")
        
        # Define the file names for the invoice, timesheet, and output files        
        invoice_file_name = f"{invoice}.pdf"
        timesheet_file_name = f"{tc_number}.pdf"

        # Determine the output file name based on the client
        if client == "Verizon - PA":
            output_file_name = f"{invoice}_{tc_number}.pdf"
        else:
            output_file_name = f"{invoice}.pdf"

        # Define the file paths for the invoice, timesheet, and output files
        invoice_file_path = os.path.join(invoice_folder, invoice_file_name).replace("\\", "/")
        timesheet_file_path = os.path.join(timesheet_folder, timesheet_file_name).replace("\\", "/")
        output_file_path = os.path.join(combined_folder, output_file_name).replace("\\", "/")

        # Append invoice file to the output file
        if os.path.exists(invoice_file_path):
            pdf_reader = PdfReader(invoice_file_path)
            for page in pdf_reader.pages:
                pdf_writer.add_page(page)

        # Append timesheet file to the output file
        if os.path.exists(timesheet_file_path):
            pdf_reader = PdfReader(timesheet_file_path)
            for page in pdf_reader.pages:
                pdf_writer.add_page(page)

        # Write the combined PDFs to a file
        try:
            with open(output_file_path, "wb") as output_file:
                pdf_writer.write(output_file)
                logger.info(f"Saving {invoice}.pdf")
        except Exception as e:
            logger.error(f"An error occurred: {e}")
            # Handle other potential errors

    logger.info("Done combining invoices!")
    print("Done combining invoices!")

def select_timesheet_folder(root):
    root.timesheet_folder = filedialog.askdirectory(title="Select Timesheet Folder")
    if root.timesheet_folder:
        root.timesheet_folder_label.config(text=f"Timesheet Folder: {os.path.basename(root.timesheet_folder)}")

def select_invoice_folder(root):
    root.invoice_folder = filedialog.askdirectory(title="Select Invoice Folder")
    if root.invoice_folder:
        root.invoice_folder_label.config(text=f"Invoice Folder: {os.path.basename(root.invoice_folder)}")

def select_combined_folder(root):
    root.combined_folder = filedialog.askdirectory(title="Select Output Folder")
    if root.combined_folder:
        root.combined_folder_label.config(text=f"Output Folder: {os.path.basename(root.combined_folder)}")

def select_excel_file(root):
    root.excel_file = filedialog.askopenfilename(title="Select the Excel Workbook", filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
    if root.excel_file:
        root.excel_file_label.config(text=f"Excel file selected: {os.path.basename(root.excel_file)}")

def start_combining(root):
    if hasattr(root, 'timesheet_folder') and hasattr(root, 'invoice_folder') and hasattr(root, 'combined_folder') and hasattr(root, 'excel_file'):
        try:
            # Load workbook and setup logger here as needed
            workbook = load_workbook(root.excel_file)
            sheet = workbook["TimesheetCombiner"]

            log_file = os.path.join(os.path.dirname(root.excel_file), "debug.log")
            logging.basicConfig(filename=log_file, level=logging.WARNING, format='%(asctime)s - %(levelname)s - %(message)s')
            logger = logging.getLogger('timesheet_combiner')

            # Here you call your combining functions with the proper arguments
            combine_timesheets(root.timesheet_folder, root.timesheet_folder, sheet, logger)
            combine_invoices(root.invoice_folder, root.timesheet_folder, root.combined_folder, sheet, logger)

            root.completion_label.config(text="Combining Completed.")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")
    else:
        messagebox.showerror("Error", "Please select all folders and the Excel file.")

def main():
    root = tk.Tk()
    root.title("Timesheet and Invoice Combiner - Verizon")
    root.geometry("400x250")

    # Timesheet Folder Selection
    root.timesheet_folder_label = tk.Label(root, text="Timesheet Folder: Not selected")
    root.timesheet_folder_label.pack()
    tk.Button(root, text="Select Timesheet Folder", command=lambda: select_timesheet_folder(root)).pack()

    # Invoice Folder Selection
    root.invoice_folder_label = tk.Label(root, text="Invoice Folder: Not selected")
    root.invoice_folder_label.pack()
    tk.Button(root, text="Select Invoice Folder", command=lambda: select_invoice_folder(root)).pack()

    # Output Folder Selection
    root.combined_folder_label = tk.Label(root, text="Combined Folder: Not selected")
    root.combined_folder_label.pack()
    tk.Button(root, text="Select Combined Folder", command=lambda: select_combined_folder(root)).pack()

    # Excel File Selection
    root.excel_file_label = tk.Label(root, text="Excel File: Not selected")
    root.excel_file_label.pack()
    tk.Button(root, text="Select Excel File", command=lambda: select_excel_file(root)).pack()

    separator = tk.Frame(root, height=2, bg="gray")
    separator.pack(fill="x", pady=10)

    # Start Button
    tk.Button(root, text="Start Combining", command=lambda: start_combining(root)).pack()

    # Completion Label
    root.completion_label = tk.Label(root, text="")
    root.completion_label.pack()

    root.mainloop()

if __name__ == "__main__":
    main()
